# backend/sales/admin.py
import io
from datetime import date, datetime, date as date_cls, time as time_cls
from django.db import transaction, models
from django.db.models import Q
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from django.contrib import admin, messages
from django.template.response import TemplateResponse
from django.urls import path
from django.shortcuts import redirect
from django.utils import timezone
from django.utils.html import format_html
from django.utils.text import Truncator
from django.core.management import call_command
from django.http import HttpResponse
import xlsxwriter
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import (
    Company, GuideProfile, BookingSale, FamilyBooking, Traveler,
    InboundEmail, CancelledBookingSale, ExcursionNetPrice
)
from .services.netto import resolve_net_prices
from .services import costasolinfo as csi
from .forms import TouristsImportForm
from .importers import tourists_excel

log = logging.getLogger(__name__)



# ── общий рендер бейджа статуса ───────────────────────────────────────────────
def render_status_badge(status: str):
    s = (status or "").lower()
    label = {
        "draft": "DRAFT",
        "pending": "Отправлено",
        "hold": "HOLD",
        "paid": "PAID",
        "cancelled": "Отменено",
        "expired": "EXPIRED",
    }.get(s, status)
    return format_html('<span class="sp-badge sp-{}">{}</span>', s, label)

# ------- фильтр «Скрыть отменённые» ------------------------------------------
class HideCancelledFilter(admin.SimpleListFilter):
    title = "Скрыть отменённые"
    parameter_name = "hide_cancelled"

    def lookups(self, request, model_admin):
        # По умолчанию скрываем, поэтому две опции:
        return (
            ("1", "Скрыть отменённые"),   # дефолт
            ("0", "Показывать все"),
        )

    def queryset(self, request, queryset):
        val = self.value()
        if val in (None, "", "1"):
            return queryset.exclude(status="CANCELLED")
        return queryset


# ───────────────────────────────────────────────────────────────────────────────
# InboundEmail
@admin.action(description="Проверить входящие Gmail сейчас")
def fetch_gmail_now(modeladmin, request, queryset):
    call_command("fetch_gmail")
    modeladmin.message_user(request, "Готово: входящие обновлены.")

@admin.register(InboundEmail)
class InboundEmailAdmin(admin.ModelAdmin):
    list_display = ("subject", "from_email", "date", "created_at")
    search_fields = ("subject", "from_email", "to_email", "snippet", "body_text")
    readonly_fields = ("uid", "message_id", "raw_headers", "created_at", "html_preview")
    actions = [fetch_gmail_now]

    def html_preview(self, obj):
        return format_html('<div style="border:1px solid #eee;padding:8px;">{}</div>', obj.body_html or "—")
    html_preview.short_description = "HTML"


# ------- вспомогалки ---------------------------------------------------------
def _status_badge(status: str) -> str:
    s = (status or "").upper()
    cls = {
        "DRAFT":     "badge badge-info",
        "PENDING":   "badge badge-warning",
        "HOLD":      "badge",
        "PAID":      "badge badge-success",
        "CONFIRMED": "badge badge-success",
        "CANCELLED": "badge badge-muted",
        "EXPIRED":   "badge badge-muted",
    }.get(s, "badge")
    return format_html('<span class="{}">{}</span>', cls, s)


def _resolve_net_for_booking(b: BookingSale):
    """
    Возвращает (currency, net_adult:Decimal, net_child:Decimal).
    Пытается использовать sales.services.netto.resolve_net_prices,
    иначе — безопасный фолбэк EUR/0/0.
    """
    try:
        if resolve_net_prices:
            reg = (getattr(b, "region_name", "") or "").strip().lower() or None
            res = resolve_net_prices(
                company_id=getattr(b.company, "id", None),
                excursion_id=int(getattr(b, "excursion_id", 0) or 0),
                region_slug=reg,
                date=getattr(b, "date", None),
            )
            if res:
                from decimal import Decimal
                return (
                    res.get("currency") or "EUR",
                    Decimal(res.get("net_adult") or 0),
                    Decimal(res.get("net_child") or 0),
                )
    except Exception:
        pass
    from decimal import Decimal
    return "EUR", Decimal("0"), Decimal("0")

@admin.action(description="Backfill region_name для выбранных продаж")
def backfill_region_name(modeladmin, request, queryset):
    fixed = 0
    for obj in queryset:
        before = (obj.region_name or "").strip()
        obj.ensure_region_name()  # метод модели
        after = (obj.region_name or "").strip()
        if after and after != before:
            obj.save(update_fields=["region_name"])
            fixed += 1
    modeladmin.message_user(request, f"Обновлено записей: {fixed}")

# ------- BookingSale (основной список продаж) --------------------------------
@admin.register(BookingSale)
class BookingSaleAdmin(admin.ModelAdmin):
    list_display = (
        "booking_code", "company", "excursion_title", "date",
        "hotel_name", "pickup_point_name", "pickup_time_str",
        "excursion_language", "room_number",
        "adults", "children", "gross_total", "status_badge",
    )
    list_filter = (HideCancelledFilter, "company", "status", "excursion_language", "date")
    search_fields = (
        "booking_code", "excursion_title", "hotel_name",
        "pickup_point_name", "room_number"
    )
    readonly_fields = ("created_at",)
    actions = ["export_bookings_xlsx", backfill_region_name]

    # цветной статус
    def status_badge(self, obj):
        return _status_badge(getattr(obj, "status", ""))
    status_badge.short_description = "Status"

    class Media:
        css = {"all": ("sales/admin-status-badges.css",)}

    # Поддержка старого параметра (?cancelled=show) — если он есть, не прячем ничего.
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.GET.get("cancelled") == "show":
            return qs
        return qs  # остальное делает HideCancelledFilter

    # Гарантируем автозаполнение region_name при сохранении из админки
    def save_model(self, request, obj, form, change):
        before = (obj.region_name or "").strip()
        # ensure_region_name() — метод модели BookingSale, который мы добавили
        obj.ensure_region_name()
        super().save_model(request, obj, form, change)
        after = (obj.region_name or "").strip()
        if not after:
            # нужен импорт наверху файла: from django.contrib import messages
            self.message_user(
                request,
                "Регион автоматически определить не удалось. Проверьте Family, настройки CSI или текст пикапа.",
                level=messages.WARNING
            )
        elif after != before:
            self.message_user(request, f"Регион определён: {after}", level=messages.INFO)

    # -------- Excel экспорт ---------------------------------------------------
    def export_bookings_xlsx(self, request, queryset):
        from datetime import datetime, date as date_cls, time as time_cls
        from decimal import Decimal
        from django.utils import timezone
        from django.http import HttpResponse  # локально, чтобы не зависеть от верхних импортов
        import openpyxl
        from openpyxl.utils import get_column_letter
        from sales.models import ExcursionNetPrice

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"

        # ----- helpers -----
        REGION_ALIASES = {
            "cds": {"cds", "costa del sol", "costa-del-sol", "costa_del_sol", "costa del-sol", "costa-del sol"},
            "malaga": {"malaga", "mlg", "má­laga", "málaga"},
            "marbella": {"marbella", "mrb"},
            "estepona": {"estepona", "est"},
        }

        def slugify_region(raw: str) -> str:
            """
            Нормализует строку региона в известный slug.
            Никаких "родителей"/догадок — только буквальная нормализация.
            """
            s = (raw or "").strip().lower()
            if not s:
                return ""
            # упрощённая очистка
            s = s.replace("_", " ").replace("-", " ").replace(",", " ")
            s = " ".join(s.split())
            # точное попадание по словарю вариантов
            for slug, variants in REGION_ALIASES.items():
                if s in variants:
                    return slug
            # простые эвристики (в пределах буквального текста)
            if "marbella" in s: return "marbella"
            if "estepona" in s: return "estepona"
            if "malaga"   in s or "málaga" in s: return "malaga"
            if "cds" in s or "costa del sol" in s: return "cds"
            return ""

        def pick_net_row_strict(booking):
            """
            Строгий выбор строки нетто:
            1) exact: company_id + region_slug
            2) fallback: company IS NULL + region_slug
            Больше НИЧЕГО. Если region_slug пуст — цены не подставляем.
            """
            ex_id = int(getattr(booking, "excursion_id", 0) or 0)
            region_slug = slugify_region(getattr(booking, "region_name", ""))
            if not ex_id or not region_slug:
                return None, region_slug  # пустой регион => нет подстановки

            base = ExcursionNetPrice.objects.filter(
                excursion_id=ex_id,
                is_active=True,
                region_slug__iexact=region_slug,
            )

            company_id = getattr(booking, "company_id", None)

            # 1) company override
            if company_id:
                row = base.filter(company_id=company_id).order_by("-updated_at", "-id").first()
                if row:
                    return row, region_slug

            # 2) general (company is NULL)
            row = base.filter(company__isnull=True).order_by("-updated_at", "-id").first()
            return row, region_slug

        def net_child(row) -> Decimal:
            if not row:
                return Decimal("0")
            # если задана детская — берём её, иначе считаем скидкой от взрослой
            try:
                if row.net_per_child not in (None, ""):
                    return Decimal(str(row.net_per_child))
                disc = Decimal(str(row.child_discount_pct or 0)) / Decimal("100")
                base = Decimal(str(row.net_per_adult or 0))
                return (base * (Decimal("1") - disc)).quantize(Decimal("0.01"))
            except Exception:
                return Decimal("0")

        def _cell(v):
            # Excel не любит tz-aware
            if isinstance(v, datetime):
                if timezone.is_aware(v):
                    v = timezone.make_naive(timezone.localtime(v))
                return v
            if isinstance(v, (date_cls, time_cls)):
                return v
            if isinstance(v, Decimal):
                return float(v)
            return v

        # ----- columns -----
        cols = [
            ("Booking code", "booking_code"),
            ("Company",     lambda o: o.company.name if o.company_id else ""),
            ("Date",        "date"),
            ("Excursion",   "excursion_title"),
            ("Hotel",       "hotel_name"),
            ("Pickup name", "pickup_point_name"),
            ("Pickup time", "pickup_time_str"),
            ("Adults",      "adults"),
            ("Children",    "children"),
            ("Gross",       "gross_total"),
            ("Language",    "excursion_language"),
            ("Room",        "room_number"),
            ("Created at",  "created_at"),
            # ↓ нетто-колонки добавим позже по месту
        ]

        # заголовки
        for i, (title, _) in enumerate(cols + [
            ("Net (adult)", None),
            ("Net (child)", None),
            ("Net total",   None),
            ("Commission",  None),
            ("Matched region", None),  # диагностика: какой slug реально применили
        ], start=1):
            ws.cell(row=1, column=i, value=title)

        # строки
        r = 2
        for obj in queryset:
            c = 1
            for _, getter in cols:
                val = getter(obj) if callable(getter) else getattr(obj, getter, "")
                ws.cell(row=r, column=c, value=_cell(val))
                c += 1

            # расчёт нетто/комиссии — строго по (excursion, company, region)
            row, used_region = pick_net_row_strict(obj)
            if row:
                try:
                    net_ad = Decimal(str(row.net_per_adult or 0))
                except Exception:
                    net_ad = Decimal("0")
                net_ch = net_child(row)
                ad = int(obj.adults or 0)
                ch = int(obj.children or 0)
                net_total = (net_ad * Decimal(ad)) + (net_ch * Decimal(ch))
                commission = (Decimal(str(obj.gross_total or 0)) - net_total)
                matched_region = used_region or "(none)"
            else:
                net_ad = net_ch = net_total = commission = Decimal("0")
                matched_region = (used_region or "(none)") + " (strict-miss)"

            ws.cell(row=r, column=c,   value=_cell(net_ad));       c += 1
            ws.cell(row=r, column=c,   value=_cell(net_ch));       c += 1
            ws.cell(row=r, column=c,   value=_cell(net_total));    c += 1
            ws.cell(row=r, column=c,   value=_cell(commission));   c += 1
            ws.cell(row=r, column=c,   value=matched_region);      c += 1  # диагностика

            r += 1

        # ширины
        for i in range(1, 1 +  len(cols) + 5):
            ws.column_dimensions[get_column_letter(i)].width = 18

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="bookings_with_net.xlsx"'
        wb.save(resp)
        return resp


# ───────────────────────────────────────────────────────────────────────────────
# Прокси-админ: только отменённые
@admin.register(CancelledBookingSale)
class CancelledBookingSaleAdmin(BookingSaleAdmin):
    # свой набор фильтров — БЕЗ HideCancelledFilter и, по желанию, без status
    list_filter = ("company", "excursion_language", "date")

    def get_queryset(self, request):
        # НЕ зовём super(), чтобы не сработала логика «скрывать CANCELLED»
        return self.model.objects.filter(status="CANCELLED")

    def changelist_view(self, request, extra_context=None):
        extra_context = (extra_context or {}) | {"title": "Cancelled bookings"}
        return super().changelist_view(request, extra_context=extra_context)


# ── Inlines для семьи ─────────────────────────────────────────────────────────
class TravelerInline(admin.TabularInline):
    model = Traveler
    extra = 0
    fields = ("last_name", "first_name", "dob", "nationality", "passport", "passport_expiry")
    show_change_link = True

class BookingSaleInline(admin.TabularInline):
    model = BookingSale
    extra = 0
    can_delete = False
    show_change_link = True

    fields = (
    "booking_code", "date",
    "excursion_title", "excursion_language",
    "adults", "children", "infants",
    "pickup_point_name", "pickup_time_str",
    "gross_total", "status_badge",
    )
    readonly_fields = fields  # показываем как read-only, редактирование в самой брони

    def status_badge(self, obj):
        return render_status_badge(getattr(obj, "status", ""))
    status_badge.short_description = "Status"

    class Media:
        css = {"all": ("sales/admin-status-badges.css",)}

# ── FamilyBooking ──────────────────────────────────────────────────────────────
@admin.register(FamilyBooking)
class FamilyBookingAdmin(admin.ModelAdmin):
    list_display = ("ref_code", "hotel_name", "arrival_date", "departure_date", "people", "created_at")
    search_fields = ("ref_code", "hotel_name", "region_name", "phone", "email",
                     "travelers__last_name", "travelers__first_name")
    list_filter = ("arrival_date", "region_name")
    inlines = [TravelerInline, BookingSaleInline]

    change_list_template = "admin/sales/familybooking/change_list.html"

    # чтобы список работал без N+1
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.prefetch_related("travelers", "bookings")

    def people(self, obj):
        # короткий список имён для списка семей
        names = ["{} {}".format(t.last_name or "", t.first_name or "").strip()
                 for t in obj.travelers.all()]
        text = ", ".join(filter(None, names)) or "—"
        return Truncator(text).chars(60)
    people.short_description = "Состав семьи"

    # импорт как у вас было
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_view),
                name="sales_familybooking_import",
            ),
        ]
        return custom + urls

    def import_view(self, request):
        if not self.has_add_permission(request):
            messages.error(request, "Недостаточно прав для импорта.")
            return redirect("admin:sales_familybooking_changelist")

        context = dict(self.admin_site.each_context(request), title="Импорт туристов")

        if request.method == "POST":
            form = TouristsImportForm(request.POST, request.FILES)
            if form.is_valid():
                up_file = form.cleaned_data.get("file")
                dry = bool(form.cleaned_data.get("dry_run", False))
                try:
                    result = tourists_excel.import_file(up_file, dry_run=dry)
                    families = result.get("families_created", 0)
                    travelers = result.get("travelers_created", 0)
                    skipped = result.get("skipped", 0)
                    msg = f"Семей: {families}, туристов: {travelers}, пропущено: {skipped}."
                    messages.success(
                        request,
                        ("Проверка прошла успешно. " if dry else "Импорт завершён. ") + msg
                    )
                    return redirect("admin:sales_familybooking_changelist")
                except Exception as e:
                    log.exception("Ошибка при импорте туристов")
                    messages.error(request, f"Ошибка импорта: {e}")
            else:
                messages.error(request, f"Проверьте форму: {form.errors.as_text()}")
        else:
            form = TouristsImportForm()

        context["form"] = form
        return TemplateResponse(request, "admin/sales/familybooking/import_form.html", context)


# ───────────────────────────────────────────────────────────────────────────────
# Traveler
@admin.register(Traveler)
class TravelerAdmin(admin.ModelAdmin):
    list_display = (
        "last_name", "first_name", "dob", "nationality", "passport",
        "passport_expiry", "gender", "doc_type", "doc_expiry", "family"
    )
    list_filter = ("gender", "doc_type", "nationality", "family")
    search_fields = ("last_name", "first_name", "passport", "email", "phone")


# ───────────────────────────────────────────────────────────────────────────────
# Прочие справочники
@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "email_for_orders", "is_active")
    search_fields = ("name", "slug", "email_for_orders")
    list_filter = ("is_active",)

@admin.register(GuideProfile)
class GuideProfileAdmin(admin.ModelAdmin):
    list_display = ("user",)
    search_fields = ("user__username", "user__email")


# ───────────────────────────────────────────────────────────────────────────────
# Цены НЕТТО по экскурсиям
@admin.register(ExcursionNetPrice)
class ExcursionNetPriceAdmin(admin.ModelAdmin):
    change_list_template = "admin/sales/excursionnetprice/change_list.html"
    list_display = ("excursion_with_title", "company", "region_slug",
                    "currency", "net_per_adult", "net_per_child",
                    "child_discount_pct", "is_active", "updated_at")
    list_filter = ("company", "region_slug", "is_active", "currency")
    search_fields = ("region_slug",)

    def excursion_with_title(self, obj):
        title = csi.excursion_title(int(obj.excursion_id or 0), lang="ru") or ""
        return f"ex#{obj.excursion_id} — {title}"
    excursion_with_title.short_description = "Excursion"

    # добавим ссылку на «Матрицу»
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("matrix/", self.admin_site.admin_view(self.matrix_view), name="sales_netprice_matrix"),
        ]
        return custom + urls

    def matrix_view(self, request):
        """
        Страница матрицы: все экскурсии × регионы.
        GET — рендер формы, POST — сохранение.
        """
        # 1) тянем все активные экскурсии из основной базы
        data = csi.list_excursions(lang="ru") or {}
        items = data.get("items") if isinstance(data, dict) else data
        excursions = []
        for it in (items or []):
            excursions.append({
                "id": int(it.get("id")),
                "title": it.get("title") or it.get("localized_title") or "",
                # Если из core у экскурсии есть «цены по регионам», берём список слагов
                "region_slugs": self._regions_for_excursion(int(it.get("id")))
            })

        # 2) справочник уже сохранённых нетто
        rows = ExcursionNetPrice.objects.all()
        saved = {}
        for r in rows:
            saved[(r.excursion_id, (r.region_slug or "").lower(), r.company_id or 0)] = r

        # 3) компании (фильтр по шапке, по желанию)
        companies = list(Company.objects.filter(is_active=True).order_by("name"))
        company_id = request.GET.get("company")
        try:
            company_id = int(company_id) if company_id else None
        except ValueError:
            company_id = None

        if request.method == "POST":
            # ожидаем поля: price-<exc_id>-<region>-adult|child
            # пример name: price-2-cds-adult
            objects_by_key: dict[tuple[int, str, int], ExcursionNetPrice] = {}

            def _parse_decimal(v: str) -> Decimal | None:
                v = (v or "").strip().replace(",", ".")
                if v == "":
                    return None
                try:
                    return Decimal(v)
                except (InvalidOperation, ValueError):
                    return None

            for k, v in request.POST.items():
                if not k.startswith("price-"):
                    continue
                try:
                    _, ex_str, reg, kind = k.split("-", 3)
                    ex_id = int(ex_str)
                except Exception:
                    continue
                reg = (reg or "").lower().strip()
                if kind not in ("adult", "child"):
                    continue

                # ключ уникальности
                comp_id = company_id or 0
                key = (ex_id, reg, comp_id)

                # берём существующую запись либо создаём новую (в памяти)
                obj = objects_by_key.get(key)
                if not obj:
                    obj = saved.get(key)
                    if not obj:
                        obj = ExcursionNetPrice(
                            excursion_id=ex_id,
                            region_slug=reg,
                            company_id=company_id,   # может быть None
                            currency="EUR",
                            is_active=True,
                        )
                    objects_by_key[key] = obj

                val = _parse_decimal(v)
                if kind == "adult":
                    obj.net_per_adult = val
                else:
                    obj.net_per_child = val

            # сохраняем всё одним проходом
            with transaction.atomic():
                for obj in objects_by_key.values():
                    obj.save()

            self.message_user(request, "Нетто-цены сохранены.")
            return redirect("admin:sales_netprice_matrix")


        context = dict(
            self.admin_site.each_context(request),
            title="Матрица нетто-цен",
            excursions=excursions,
            companies=companies,
            sel_company=company_id,
            saved=saved,
        )
        return TemplateResponse(request, "admin/sales/net_prices_matrix.html", context)

    def _regions_for_excursion(self, excursion_id: int) -> list[str]:
        """Пробуем вытащить список регионов из детальной инфы экскурсии.
        Если нет — вернём базовый набор."""
        detail = csi.excursion_detail(excursion_id) or {}
        regions = set()
        for key in ("prices_by_region", "pricesByRegion", "region_prices", "prices", "tariffs"):
            arr = detail.get(key) or []
            if isinstance(arr, list):
                for row in arr:
                    r = (row.get("region") or {})
                    slug = (r.get("slug") or r.get("code") or "").lower()
                    if slug:
                        regions.add(slug)
        # бэкап-набор, если источник молчит:
        if not regions:
            regions = {"cds", "malaga", "marbella", "estepona"}
        return sorted(regions)